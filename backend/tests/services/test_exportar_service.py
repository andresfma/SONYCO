import io
import pytest
from openpyxl import load_workbook

from app.services.exportar_service import (
    exportar_clientes,
    exportar_categorias,
    exportar_productos,
    exportar_usuarios,
    exportar_inventario,
    exportar_ventas,
    exportar_ventas_por_cliente,
    exportar_movimientos_por_producto,
    exportar_movimientos_por_usuario,
    exportar_detalle_venta_por_venta_id,
    exportar_movimientos_inventario,
)


# -------------------------------
# Helpers
# -------------------------------

def _load_excel(bytes_io: io.BytesIO):
    """Carga un Excel de BytesIO y devuelve workbook y worksheet activo"""
    wb = load_workbook(bytes_io)
    return wb, wb.active

def _get_headers(ws):
    """Obtiene los headers de la primera fila"""
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]


# -------------------------------
# Clientes
# -------------------------------

def test_exportar_clientes(session, cliente_fixture):
    excel_bytes = exportar_clientes(session)

    # 1. Retorno BytesIO
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    # 2. Headers correctos
    headers = _get_headers(ws)
    assert headers == ["ID", "Identificación", "Nombre", "Tipo", "Email", "Teléfono", "Estado"]

    # 3. Datos correctos
    row = [cell.value for cell in ws[2]]
    assert cliente_fixture.identificacion in row
    assert cliente_fixture.nombre in row


# -------------------------------
# Categorías
# -------------------------------

def test_exportar_categorias(session, categoria_fixture):
    excel_bytes = exportar_categorias(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    headers = _get_headers(ws)
    assert headers == ["ID", "Nombre", "Descripción", "Estado"]

    row = [cell.value for cell in ws[2]]
    assert categoria_fixture.nombre in row
    assert categoria_fixture.descripcion in row


# -------------------------------
# Productos
# -------------------------------

def test_exportar_productos(session, producto_fixture):
    excel_bytes = exportar_productos(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    headers = _get_headers(ws)
    assert headers == ["ID", "Código", "Nombre", "Unidad_medida", "Descripción", "Precio Unitario", "Categoria", "Estado"]

    row = [cell.value for cell in ws[2]]
    assert producto_fixture.codigo in row
    assert producto_fixture.nombre in row


# -------------------------------
# Usuarios
# -------------------------------

def test_exportar_usuarios(session, usuario_fixture):
    excel_bytes = exportar_usuarios(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    headers = _get_headers(ws)
    assert headers == ["ID", "Nombre", "Email", "Rol", "Estado"]

    row = [cell.value for cell in ws[2]]
    assert usuario_fixture.nombre in row
    assert usuario_fixture.email in row


# -------------------------------
# Inventario
# -------------------------------

def test_exportar_inventario(session, inventario_fixture):
    excel_bytes = exportar_inventario(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    headers = _get_headers(ws)
    assert headers == ["ID", "Código_producto", "Producto", "Unidad_medida", "Cantidad", "Cantidad_minima", "Precio", "Categoria", "Estado"]

    row = [cell.value for cell in ws[2]]
    assert inventario_fixture.producto.nombre in row


# -------------------------------
# Ventas
# -------------------------------

def test_exportar_ventas(session, venta_fixture):
    excel_bytes = exportar_ventas(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)

    headers = _get_headers(ws)
    assert headers == ["ID", "Identificación_cliente", "Cliente", "Vendedor", "Fecha", "Total", "Estado"]

    row = [cell.value for cell in ws[2]]
    assert venta_fixture.cliente.nombre in row
    assert venta_fixture.usuario.nombre in row


def test_exportar_ventas_por_cliente(session, venta_fixture):
    excel_bytes = exportar_ventas_por_cliente(session, venta_fixture.cliente_id)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    assert headers == ["ID", "Identificación_cliente", "Cliente", "Estado_cliente", "Vendedor", "Fecha", "Total", "Estado_venta"]

    row = [cell.value for cell in ws[2]]
    assert venta_fixture.cliente.nombre in row
    assert venta_fixture.usuario.nombre in row


# -------------------------------
# Movimientos
# -------------------------------

def test_exportar_movimientos_por_producto(session, movimiento_fixture):
    excel_bytes = exportar_movimientos_por_producto(session, movimiento_fixture.producto_id)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    assert "Código_producto" in headers

    row = [cell.value for cell in ws[2]]
    assert movimiento_fixture.producto.nombre in row


def test_exportar_movimientos_por_usuario(session, movimiento_fixture):
    excel_bytes = exportar_movimientos_por_usuario(session, movimiento_fixture.usuario_id)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    assert "Empleado" in headers

    row = [cell.value for cell in ws[2]]
    assert movimiento_fixture.usuario.nombre in row


def test_exportar_movimientos_inventario(session, movimiento_fixture):
    excel_bytes = exportar_movimientos_inventario(session)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    assert "Tipo Movimiento" in headers

    row = [cell.value for cell in ws[2]]
    assert movimiento_fixture.producto.nombre in row


# -------------------------------
# Detalle Venta
# -------------------------------

def test_exportar_detalle_venta_por_venta_id(session, detalle_venta_fixture):
    excel_bytes = exportar_detalle_venta_por_venta_id(session, detalle_venta_fixture.venta_id)
    assert isinstance(excel_bytes, io.BytesIO)

    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    assert headers == ["ID", "Código_producto", "Producto", "Unidad_medida", "Estado_producto", "Cantidad", "Precio Unitario", "Subtotal"]

    row = [cell.value for cell in ws[2]]
    assert detalle_venta_fixture.producto.nombre in row
