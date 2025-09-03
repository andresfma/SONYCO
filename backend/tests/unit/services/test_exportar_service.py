import io
from openpyxl import load_workbook

from app.services.exportar_service import (
    exportar_clientes,
    exportar_categorias,
    exportar_productos,
    exportar_usuarios,
    exportar_inventario,
    exportar_ventas,
    exportar_ventas_por_cliente,
    exportar_movimientos_por_producto,
    exportar_movimientos_por_usuario,
    exportar_detalle_venta_por_venta_id,
    exportar_movimientos_inventario,
)
from app.models.cliente import Cliente, TipoPersona
from app.models.categoria import  Categoria
from app.models.producto import Producto, UnidadMedida



# -------------------------------
# Helpers
# -------------------------------

def _load_excel(bytes_io: io.BytesIO):
    """Carga un Excel de BytesIO y devuelve workbook y worksheet activo"""
    wb = load_workbook(bytes_io)
    return wb, wb.active

def _get_headers(ws):
    """Obtiene los headers de la primera fila"""
    return [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]

def _get_row_count(ws):
    """Obtiene el número de filas con datos (excluyendo headers)"""
    return ws.max_row - 1

def _validate_excel_structure(excel_bytes, expected_headers):
    """Valida la estructura básica del Excel"""
    assert isinstance(excel_bytes, io.BytesIO)
    assert excel_bytes.getvalue(), "El archivo Excel está vacío"
    
    wb, ws = _load_excel(excel_bytes)
    headers = _get_headers(ws)
    
    assert headers == expected_headers, f"Headers incorrectos. Esperados: {expected_headers}, Obtenidos: {headers}"
    
    return wb, ws


# -------------------------------
# Tests para Clientes
# -------------------------------

class TestExportarClientes:
    """Tests para exportación de clientes"""
    
    EXPECTED_HEADERS = ["ID", "Identificación", "Nombre", "Tipo", "Email", "Teléfono", "Estado"]
    
    def test_exportar_clientes_con_datos(self, session, cliente_fixture):
        """Test exportar clientes con datos existentes"""
        excel_bytes = exportar_clientes(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que hay datos (al menos 1 fila además del header)
        assert _get_row_count(ws) >= 1
        
        # Verificar datos del cliente fixture
        row_data = [cell.value for cell in ws[2]]
        assert cliente_fixture.id in row_data
        assert cliente_fixture.identificacion in row_data
        assert cliente_fixture.nombre in row_data
        assert cliente_fixture.tipo_persona.value in row_data
        assert cliente_fixture.email in row_data
        assert cliente_fixture.telefono in row_data
    
    def test_exportar_clientes_multiples(self, session, clientes_fixture):
        """Test exportar múltiples clientes"""
        excel_bytes = exportar_clientes(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar cantidad de registros
        assert _get_row_count(ws) == len(clientes_fixture)
        
        # Verificar que todos los clientes están presentes
        nombres_en_excel = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            nombres_en_excel.append(row[2])  # Columna "Nombre"
        
        for cliente in clientes_fixture:
            assert cliente.nombre in nombres_en_excel
    
    def test_exportar_clientes_sin_datos(self, session):
        """Test exportar clientes cuando no hay datos"""
        excel_bytes = exportar_clientes(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Solo debe tener el header
        assert _get_row_count(ws) == 0
    
    def test_exportar_clientes_diferentes_tipos_persona(self, session):
        """Test exportar clientes con diferentes tipos de persona"""
        # Crear clientes de diferentes tipos
        cliente_natural = Cliente(
            nombre="Persona Natural",
            email="natural@test.com",
            tipo_persona=TipoPersona.natural,
            identificacion="1111111111"
        )
        cliente_juridica = Cliente(
            nombre="Empresa S.A.",
            email="empresa@test.com",
            tipo_persona=TipoPersona.juridica,
            identificacion="J22222222"
        )
        
        session.add_all([cliente_natural, cliente_juridica])
        session.commit()
        
        excel_bytes = exportar_clientes(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar tipos de persona en el Excel
        tipos_persona = []
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            tipos_persona.append(row[3])  # Columna "Tipo"
        
        assert "natural" in tipos_persona
        assert "juridica" in tipos_persona


# -------------------------------
# Tests para Categorías
# -------------------------------

class TestExportarCategorias:
    """Tests para exportación de categorías"""
    
    EXPECTED_HEADERS = ["ID", "Nombre", "Descripción", "Estado"]
    
    def test_exportar_categorias_con_datos(self, session, categoria_fixture):
        """Test exportar categorías con datos existentes"""
        excel_bytes = exportar_categorias(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert categoria_fixture.id in row_data
        assert categoria_fixture.nombre in row_data
        assert categoria_fixture.descripcion in row_data
    
    def test_exportar_categorias_sin_descripcion(self, session):
        """Test exportar categorías con descripción None"""
        categoria_sin_desc = Categoria(
            nombre="Categoría Sin Descripción",
            descripcion=None
        )
        session.add(categoria_sin_desc)
        session.commit()
        
        excel_bytes = exportar_categorias(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que maneja None correctamente
        row_data = [cell.value for cell in ws[2]]
        assert "Categoría Sin Descripción" in row_data
        assert None in row_data or "" in row_data  # Dependiendo de cómo maneja openpyxl los None
    
    def test_exportar_categorias_estados_mixtos(self, session, categoria_fixture):
        """Test exportar categorías con estados activo/inactivo"""
        categoria_inactiva = Categoria(
            nombre="Categoría Inactiva",
            descripcion="Descripción inactiva",
            estado=False
        )
        session.add(categoria_inactiva)
        session.commit()
        
        excel_bytes = exportar_categorias(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 2


# -------------------------------
# Tests para Productos
# -------------------------------

class TestExportarProductos:
    """Tests para exportación de productos"""
    
    EXPECTED_HEADERS = ["ID", "Código", "Nombre", "Unidad_medida", "Descripción", "Precio Unitario", "Categoria", "Estado"]
    
    def test_exportar_productos_con_datos(self, session, producto_fixture):
        """Test exportar productos con datos existentes"""
        excel_bytes = exportar_productos(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert producto_fixture.codigo in row_data
        assert producto_fixture.nombre in row_data
        assert producto_fixture.precio_unitario in row_data
    
    def test_exportar_productos_con_categoria(self, session, producto_fixture):
        """Test exportar productos verificando relación con categoría"""
        excel_bytes = exportar_productos(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        row_data = [cell.value for cell in ws[2]]
        if producto_fixture.categoria:
            assert producto_fixture.categoria.nombre in row_data
    
    def test_exportar_productos_precios_decimales(self, session, categoria_fixture):
        """Test exportar productos con precios decimales"""
        producto_decimal = Producto(
            codigo="PROD001",
            nombre="Producto Decimal",
            precio_unitario=15.99,
            unidad_medida=UnidadMedida.KILOGRAMO,
            categoria_id=categoria_fixture.id
        )
        session.add(producto_decimal)
        session.commit()
        
        excel_bytes = exportar_productos(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que los decimales se mantienen
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[1] == "PROD001":  # Columna código
                assert row[5] == 15.99  # Columna precio


# -------------------------------
# Tests para Usuarios
# -------------------------------

class TestExportarUsuarios:
    """Tests para exportación de usuarios"""
    
    EXPECTED_HEADERS = ["ID", "Nombre", "Email", "Rol", "Estado"]
    
    def test_exportar_usuarios_con_datos(self, session, usuario_fixture):
        """Test exportar usuarios con datos existentes"""
        excel_bytes = exportar_usuarios(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert usuario_fixture.nombre in row_data
        assert usuario_fixture.email in row_data

        if usuario_fixture.rol_id == 1:
            rol_usuario = "Admin"
        else:
            rol_usuario = "No-admin"

        assert rol_usuario in row_data



# -------------------------------
# Tests para Inventario
# -------------------------------

class TestExportarInventario:
    """Tests para exportación de inventario"""
    
    EXPECTED_HEADERS = ["ID", "Código_producto", "Producto", "Unidad_medida", "Cantidad", "Cantidad_minima", "Precio", "Categoria", "Estado"]
    
    def test_exportar_inventario_con_datos(self, session, inventario_fixture):
        """Test exportar inventario con datos existentes"""
        excel_bytes = exportar_inventario(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert inventario_fixture.producto.nombre in row_data
        assert inventario_fixture.cantidad in row_data
        assert inventario_fixture.cantidad_minima in row_data
    
    def test_exportar_inventario_stock_bajo(self, session, inventario_fixture):
        """Test exportar inventario verificando stock bajo"""
        # Modificar inventario para que tenga stock bajo
        inventario_fixture.cantidad = 2
        inventario_fixture.cantidad_minima = 10
        session.commit()
        
        excel_bytes = exportar_inventario(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que los valores se exportan correctamente
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] == inventario_fixture.id:  # ID del inventario
                assert row[4] == 2   # Cantidad
                assert row[5] == 10  # Cantidad mínima


# -------------------------------
# Tests para Ventas
# -------------------------------

class TestExportarVentas:
    """Tests para exportación de ventas"""
    
    EXPECTED_HEADERS = ["ID", "Identificación_cliente", "Cliente", "Vendedor", "Fecha", "Total", "Estado"]
    
    def test_exportar_ventas_con_datos(self, session, venta_fixture):
        """Test exportar ventas con datos existentes"""
        excel_bytes = exportar_ventas(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert venta_fixture.cliente.nombre in row_data
        assert venta_fixture.usuario.nombre in row_data
        assert venta_fixture.total in row_data
    
    def test_exportar_ventas_por_cliente_con_datos(self, session, venta_fixture):
        """Test exportar ventas por cliente específico"""
        expected_headers = ["ID", "Identificación_cliente", "Cliente", "Estado_cliente", "Vendedor", "Fecha", "Total", "Estado_venta"]
        
        excel_bytes = exportar_ventas_por_cliente(session, venta_fixture.cliente_id)
        wb, ws = _validate_excel_structure(excel_bytes, expected_headers)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert venta_fixture.cliente.nombre in row_data
        assert venta_fixture.usuario.nombre in row_data
    
    def test_exportar_ventas_por_cliente_sin_ventas(self, session, cliente_fixture):
        """Test exportar ventas por cliente sin ventas"""
        expected_headers = ["ID", "Identificación_cliente", "Cliente", "Estado_cliente", "Vendedor", "Fecha", "Total", "Estado_venta"]
        
        excel_bytes = exportar_ventas_por_cliente(session, cliente_fixture.id)
        wb, ws = _validate_excel_structure(excel_bytes, expected_headers)
        
        # No debe haber datos (solo headers)
        assert _get_row_count(ws) == 0
    
    def test_exportar_ventas_formato_fecha(self, session, venta_fixture):
        """Test verificar formato de fecha en exportación"""
        excel_bytes = exportar_ventas(session)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que la fecha se exporta correctamente
        fecha_cell = ws.cell(row=2, column=5).value  # Columna "Fecha"
        assert fecha_cell is not None


# -------------------------------
# Tests para Movimientos
# -------------------------------

class TestExportarMovimientos:
    """Tests para exportación de movimientos"""
    
    def test_exportar_movimientos_por_producto(self, session, movimiento_fixture):
        """Test exportar movimientos por producto"""
        excel_bytes = exportar_movimientos_por_producto(session, movimiento_fixture.producto_id)
        
        assert isinstance(excel_bytes, io.BytesIO)
        wb, ws = _load_excel(excel_bytes)
        headers = _get_headers(ws)
        
        assert "Código_producto" in headers
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert movimiento_fixture.producto.nombre in row_data
    
    def test_exportar_movimientos_por_usuario(self, session, movimiento_fixture):
        """Test exportar movimientos por usuario"""
        excel_bytes = exportar_movimientos_por_usuario(session, movimiento_fixture.usuario_id)
        
        assert isinstance(excel_bytes, io.BytesIO)
        wb, ws = _load_excel(excel_bytes)
        headers = _get_headers(ws)
        
        assert "Empleado" in headers
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert movimiento_fixture.usuario.nombre in row_data
    
    def test_exportar_movimientos_inventario_general(self, session, movimiento_fixture):
        """Test exportar todos los movimientos de inventario"""
        excel_bytes = exportar_movimientos_inventario(session)
        
        assert isinstance(excel_bytes, io.BytesIO)
        wb, ws = _load_excel(excel_bytes)
        headers = _get_headers(ws)
        
        assert "Tipo Movimiento" in headers
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert movimiento_fixture.producto.nombre in row_data
    
    def test_exportar_movimientos_producto_sin_movimientos(self, session, producto_fixture):
        """Test exportar movimientos de producto sin movimientos"""
        excel_bytes = exportar_movimientos_por_producto(session, producto_fixture.id)
        
        assert isinstance(excel_bytes, io.BytesIO)
        wb, ws = _load_excel(excel_bytes)
        
        # Debe tener headers pero no datos
        assert _get_row_count(ws) == 0


# -------------------------------
# Tests para Detalle de Venta
# -------------------------------

class TestExportarDetalleVenta:
    """Tests para exportación de detalle de venta"""
    
    EXPECTED_HEADERS = ["ID", "Código_producto", "Producto", "Unidad_medida", "Estado_producto", "Cantidad", "Precio Unitario", "Subtotal"]
    
    def test_exportar_detalle_venta_con_datos(self, session, detalle_venta_fixture):
        """Test exportar detalle de venta con datos"""
        excel_bytes = exportar_detalle_venta_por_venta_id(session, detalle_venta_fixture.venta_id)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        assert _get_row_count(ws) >= 1
        
        row_data = [cell.value for cell in ws[2]]
        assert detalle_venta_fixture.producto.nombre in row_data
        assert detalle_venta_fixture.cantidad in row_data
        assert detalle_venta_fixture.precio_unitario in row_data
    
    def test_exportar_detalle_venta_calculo_subtotal(self, session, detalle_venta_fixture):
        """Test verificar cálculo de subtotal en exportación"""
        excel_bytes = exportar_detalle_venta_por_venta_id(session, detalle_venta_fixture.venta_id)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # Verificar que subtotal = cantidad * precio_unitario
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if row[0] == detalle_venta_fixture.id:  # ID del detalle
                cantidad = row[5]      # Columna "Cantidad"
                precio = row[6]        # Columna "Precio Unitario"
                subtotal = row[7]      # Columna "Subtotal"
                
                assert abs(subtotal - (cantidad * precio)) < 0.01  # Comparación con tolerancia para decimales
    
    def test_exportar_detalle_venta_sin_detalles(self, session, venta_fixture):
        """Test exportar detalle de venta sin detalles"""
        excel_bytes = exportar_detalle_venta_por_venta_id(session, venta_fixture.id)
        wb, ws = _validate_excel_structure(excel_bytes, self.EXPECTED_HEADERS)
        
        # No debe haber datos si la venta no tiene detalles
        assert _get_row_count(ws) == 0


# -------------------------------
# Tests Generales de Integridad
# -------------------------------

class TestIntegridadExportacion:
    """Tests generales para verificar integridad de las exportaciones"""
    
    def test_encoding_caracteres_especiales(self, session):
        """Test manejo de caracteres especiales en exportación"""
        
        cliente_especial = Cliente(
            nombre="José María Ñoño",
            email="jose@test.com",
            tipo_persona=TipoPersona.natural,
            identificacion="1234567890",
            telefono="123-456-789",
            direccion="Calle 123 #45-67"
        )
        session.add(cliente_especial)
        session.commit()
        
        excel_bytes = exportar_clientes(session)
        wb, ws = _load_excel(excel_bytes)
        
        # Verificar que los caracteres especiales se mantienen
        found_special_chars = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, values_only=True):
            if "José María Ñoño" in str(row):
                found_special_chars = True
                break
        
        assert found_special_chars
    
    def test_excel_file_size_reasonable(self, session, clientes_fixture):
        """Test que el tamaño del archivo Excel sea razonable"""
        excel_bytes = exportar_clientes(session)
        
        file_size = len(excel_bytes.getvalue())
        
        # El archivo no debe estar vacío ni ser excesivamente grande
        assert file_size > 1000, "El archivo es muy pequeño, podría estar corrupto"
        assert file_size < 10_000_000, "El archivo es muy grande para pocos datos"
    
    def test_excel_workbook_metadata(self, session, cliente_fixture):
        """Test metadatos del workbook Excel"""
        excel_bytes = exportar_clientes(session)
        wb, ws = _load_excel(excel_bytes)
        
        # Verificar que el workbook tiene al menos una hoja
        assert len(wb.worksheets) >= 1
        
        # Verificar que la hoja activa tiene el formato esperado
        assert ws.max_column >= 7  # Número de columnas esperadas para clientes
        assert ws.max_row >= 1     # Al menos el header